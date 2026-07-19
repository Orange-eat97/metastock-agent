from __future__ import annotations

import json
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


DEFAULT_EXCEL_PATH = Path("data") / "explorer_outputs.xlsx"
SHEET_NAME = "explorers"


HEADERS = [
    "created_at",
    "backend",
    "model",
    "user_query",
    "explorer_name",
    "explorer_description",
    "explorer_code_body",
    "col_definitions_json",
    "validation_passed",
    "validation_errors_json",
    "full_output_json",
]


def _ensure_parent_dir(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)


def _get_or_create_sheet(path: Path) -> tuple[Workbook, Worksheet]:
    _ensure_parent_dir(path)

    if path.exists():
        wb = load_workbook(path)
        if SHEET_NAME in wb.sheetnames:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.create_sheet(SHEET_NAME)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = SHEET_NAME

    if ws.max_row == 1 and ws.cell(row=1, column=1).value is None:
        ws.append(HEADERS)

    return wb, ws


def _json_dumps(value: Any) -> str:
    return json.dumps(value, ensure_ascii=False)


def save_explorer_output_to_excel(
    *,
    output: dict[str, Any],
    user_query: str,
    backend: str,
    model: str,
    validation_errors: list[str] | None = None,
    excel_path: str | Path = DEFAULT_EXCEL_PATH,
) -> Path:
    """
    Append one generated Explorer JSON object to a local Excel sheet.

    This is intentionally row-based so it can later map cleanly to Supabase.
    Complex fields are stored as JSON strings.
    """
    path = Path(excel_path)
    wb, ws = _get_or_create_sheet(path)

    validation_errors = validation_errors or []
    validation_passed = not validation_errors

    row = [
        datetime.now().isoformat(timespec="seconds"),
        backend,
        model,
        user_query,
        output.get("explorer_name", ""),
        output.get("explorer_description", ""),
        output.get("explorer_code_body", ""),
        _json_dumps(output.get("col_definitions", [])),
        validation_passed,
        _json_dumps(validation_errors),
        _json_dumps(output),
    ]

    ws.append(row)

    # Basic readability formatting.
    for col in range(1, len(HEADERS) + 1):
        ws.cell(row=1, column=col).style = "Headline 3"

    widths = {
        "A": 22,
        "B": 14,
        "C": 24,
        "D": 45,
        "E": 28,
        "F": 45,
        "G": 45,
        "H": 45,
        "I": 40,
        "J": 55,
        "K": 35,
        "L": 18,
        "M": 45,
        "N": 60,
    }

    for col, width in widths.items():
        ws.column_dimensions[col].width = width

    wb.save(path)
    return path